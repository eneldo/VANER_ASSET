import base64
import unittest
from datetime import datetime
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from types import SimpleNamespace
from unittest.mock import patch
from uuid import uuid4

from fastapi import HTTPException
from openpyxl import load_workbook

from app.routers.reportes import (
    _crear_excel_consolidado,
    _crear_excel_individual,
    _empresa_permitida,
    _serializar_mantenimiento,
)


class ReportesProTests(unittest.TestCase):
    def setUp(self):
        self.empresa_id = uuid4()
        self.sede_id = uuid4()
        self.equipo_id = uuid4()
        self.mantenimiento_id = uuid4()
        self.equipo = SimpleNamespace(
            id=self.equipo_id,
            empresa_id=self.empresa_id,
            sede_id=self.sede_id,
            nombre="Mini split",
            inventario="INV-015",
            codigo_id="EQ-015",
            serie="SERIE-15",
            marca="Marca",
            modelo="Modelo",
            ubicacion="Consultorio medico 15",
        )
        self.empresa = SimpleNamespace(id=self.empresa_id, nombre="Clinica Ejemplo")
        self.sede = SimpleNamespace(id=self.sede_id, nombre="Sede Principal")
        self.tecnico = SimpleNamespace(
            usuario=SimpleNamespace(nombre_completo="Tecnico Prueba")
        )
        self.mantenimiento = SimpleNamespace(
            id=self.mantenimiento_id,
            equipo_id=self.equipo_id,
            empresa_id=self.empresa_id,
            sede_id=self.sede_id,
            equipo=self.equipo,
            empresa=self.empresa,
            sede=self.sede,
            tecnico=self.tecnico,
            tipo="PREVENTIVO",
            estado="FINALIZADO",
            fecha_programada=datetime(2026, 8, 10, 8, 0),
            fecha_inicio=datetime(2026, 8, 10, 9, 0),
            fecha_fin=None,
            fecha_finalizacion=datetime(2026, 8, 10, 11, 0),
            costo=0,
            descripcion="Mantenimiento programado",
            estado_inicial=None,
            estado_inicial_equipo=None,
            acciones_realizadas=None,
            resultado_final=None,
            observaciones="Equipo operativo",
        )
        self.evidencias = [
            SimpleNamespace(tipo="ANTES", descripcion="Filtro con suciedad"),
            SimpleNamespace(tipo="DURANTE", descripcion="Limpieza y ajuste electrico"),
            SimpleNamespace(tipo="DESPUES", descripcion="Equipo probado y operativo"),
        ]

    def test_coordinador_solo_consulta_su_empresa(self):
        coordinador = SimpleNamespace(rol="COORDINADOR", empresa_id=self.empresa_id)

        self.assertEqual(_empresa_permitida(coordinador), self.empresa_id)
        self.assertEqual(
            _empresa_permitida(coordinador, self.empresa_id), self.empresa_id
        )
        with self.assertRaises(HTTPException) as contexto:
            _empresa_permitida(coordinador, uuid4())
        self.assertEqual(contexto.exception.status_code, 403)

    def test_admin_puede_consultar_scope_global_o_empresa(self):
        admin = SimpleNamespace(rol="ADMIN", empresa_id=None)

        self.assertIsNone(_empresa_permitida(admin))
        self.assertEqual(
            _empresa_permitida(admin, self.empresa_id), self.empresa_id
        )

    def test_informe_usa_inventario_y_descripciones_de_evidencias(self):
        detalle = _serializar_mantenimiento(
            self.mantenimiento, SimpleNamespace(), self.evidencias
        )

        self.assertEqual(detalle["empresa"], "Clinica Ejemplo")
        self.assertEqual(detalle["sede"], "Sede Principal")
        self.assertEqual(detalle["equipo"], "Mini split")
        self.assertEqual(detalle["codigo_inventario"], "INV-015")
        self.assertEqual(detalle["ubicacion"], "Consultorio medico 15")
        self.assertEqual(detalle["tecnico"], "Tecnico Prueba")
        self.assertEqual(detalle["estado_inicial"], "Filtro con suciedad")
        self.assertEqual(
            detalle["acciones_realizadas"], "Limpieza y ajuste electrico"
        )
        self.assertEqual(detalle["resultado_final"], "Equipo probado y operativo")
        self.assertEqual(detalle["evidencias_total"], 3)

    def test_excel_consolidado_contiene_datos_tecnicos(self):
        detalle = _serializar_mantenimiento(
            self.mantenimiento, SimpleNamespace(), self.evidencias
        )
        workbook = load_workbook(BytesIO(_crear_excel_consolidado([detalle]).getvalue()))
        sheet = workbook["Mantenimientos"]

        self.assertEqual(sheet["A1"].value, "REPORTE PRO DE MANTENIMIENTOS")
        self.assertEqual(sheet["B5"].value, "Clinica Ejemplo")
        self.assertEqual(sheet["C5"].value, "Sede Principal")
        self.assertEqual(sheet["E5"].value, "INV-015")
        self.assertEqual(sheet["F5"].value, "Consultorio medico 15")
        self.assertEqual(sheet["N5"].value, "Limpieza y ajuste electrico")
        self.assertEqual(sheet["Q5"].value, 3)

    def test_excel_individual_incluye_hoja_evidencias_e_imagen(self):
        detalle = _serializar_mantenimiento(
            self.mantenimiento, SimpleNamespace(), self.evidencias
        )
        png = base64.b64decode(
            "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mP8/x8AAusB9Y9ZQmcAAAAASUVORK5CYII="
        )
        with TemporaryDirectory() as temp_dir:
            image_path = Path(temp_dir) / "antes.png"
            image_path.write_bytes(png)
            evidencia = SimpleNamespace(
                tipo="ANTES",
                descripcion="Estado inicial",
                nombre_original="antes.png",
                archivo_url="evidencias/antes.png",
            )
            with patch(
                "app.routers.reportes.get_evidencia_path", return_value=image_path
            ):
                output = _crear_excel_individual(detalle, [evidencia])

        workbook = load_workbook(BytesIO(output.getvalue()))
        self.assertIn("Informe OT", workbook.sheetnames)
        self.assertIn("Evidencias", workbook.sheetnames)
        evidence_sheet = workbook["Evidencias"]
        self.assertEqual(evidence_sheet["A2"].value, "ANTES")
        self.assertEqual(evidence_sheet["B2"].value, "Estado inicial")
        self.assertEqual(len(evidence_sheet._images), 1)


if __name__ == "__main__":
    unittest.main()
