import unittest
from datetime import datetime

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.routers.equipos import COLUMNAS_EXPORTACION, crear_excel_inventario


class EquiposExportacionTests(unittest.TestCase):
    def test_excel_conserva_columnas_y_datos_relacionados(self):
        contenido = crear_excel_inventario([
            {
                "codigo_inventario": "EQ-001",
                "nombre": "Aire consultorio",
                "empresa": "ESE Salud Yopal",
                "sede": "Hospital Central de Yopal",
                "categoria": "Aires Acondicionados",
                "marca": "=MARCA_PELIGROSA",
                "modelo": "M-100",
                "serie": None,
                "ubicacion": "Consultorio 1",
                "estado": "OPERATIVO",
                "criticidad": "MEDIA",
                "invima": None,
                "inventario": "INV-01",
                "activo": "SI",
                "fecha_creacion": datetime(2026, 8, 12, 10, 30),
            }
        ])

        workbook = load_workbook(contenido, data_only=False)
        sheet = workbook["Inventario"]

        self.assertEqual(
            [sheet.cell(1, column).value for column in range(1, 16)],
            COLUMNAS_EXPORTACION,
        )
        self.assertEqual(sheet["A2"].value, "EQ-001")
        self.assertEqual(sheet["C2"].value, "ESE Salud Yopal")
        self.assertEqual(sheet["D2"].value, "Hospital Central de Yopal")
        self.assertEqual(sheet["F2"].value, "'=MARCA_PELIGROSA")
        self.assertEqual(sheet["H2"].value, "SIN DATO")
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(len(sheet.data_validations.dataValidation), 2)

    def test_excel_usa_inventario_si_codigo_id_esta_vacio(self):
        contenido = crear_excel_inventario([
            {
                "codigo_inventario": None,
                "inventario": "12147",
                "nombre": "Mini Split",
            },
            {
                "codigo_inventario": "   ",
                "inventario": "  12148  ",
                "nombre": "Mini Split",
            },
            {
                "codigo_inventario": None,
                "inventario": None,
                "nombre": "Mini Split",
            },
        ])

        workbook = load_workbook(contenido, data_only=False)
        sheet = workbook["Inventario"]

        self.assertEqual(sheet["A2"].value, "12147")
        self.assertEqual(sheet["A3"].value, "12148")
        self.assertEqual(sheet["A4"].value, "SIN-INVENTARIO-0004")

    def test_ruta_exportar_requiere_autenticacion(self):
        response = TestClient(app).get("/equipos/exportar")
        self.assertEqual(response.status_code, 401)


if __name__ == "__main__":
    unittest.main()
