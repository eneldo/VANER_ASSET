# ============================================================
# UTILIDAD: Exportador Excel PRO
# Proyecto: SGA PRO - Fase 27 Exportación PRO
# Archivo: backend/app/utils/excel_exporter.py
#
# Función:
# - Crear archivos Excel .xlsx profesionales desde listas de datos.
# - Aplicar encabezados, filtros, congelar primera fila y ancho automático.
# - Evitar lógica repetida en el router de exportaciones.
# ============================================================

from datetime import datetime
from pathlib import Path
from typing import Iterable, List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Carpeta base donde se guardan los Excel generados.
BASE_DIR = Path(__file__).resolve().parents[1]
EXPORT_DIR = BASE_DIR / "exports" / "excel"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)


def _safe_value(value: Any) -> Any:
    """
    Convierte valores complejos a formatos seguros para Excel.
    Excel no recibe bien objetos UUID/datetime no serializados en algunos casos.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value) if not isinstance(value, (int, float)) else value


def crear_excel(nombre_base: str, titulo: str, columnas: List[str], filas: Iterable[Dict[str, Any]]) -> Path:
    """
    Crea un archivo Excel con formato institucional.

    Parámetros:
    - nombre_base: nombre técnico del archivo sin extensión.
    - titulo: título visible en la hoja.
    - columnas: columnas/keys esperadas en cada fila.
    - filas: iterable de diccionarios con la información exportada.

    Retorna:
    - Path absoluto del archivo creado.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archivo = EXPORT_DIR / f"{nombre_base}_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte SGA"

    # ========================================================
    # TÍTULO PRINCIPAL
    # ========================================================
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columnas), 1))
    celda_titulo = ws.cell(row=1, column=1, value=titulo)
    celda_titulo.font = Font(bold=True, size=16, color="FFFFFF")
    celda_titulo.fill = PatternFill("solid", fgColor="0F4C81")
    celda_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Fecha de generación.
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(columnas), 1))
    celda_fecha = ws.cell(row=2, column=1, value=f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    celda_fecha.font = Font(italic=True, color="64748B")
    celda_fecha.alignment = Alignment(horizontal="center")

    # ========================================================
    # ENCABEZADOS
    # ========================================================
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CBD5E1")

    for col_index, columna in enumerate(columnas, start=1):
        cell = ws.cell(row=4, column=col_index, value=columna.replace("_", " ").upper())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # ========================================================
    # DATOS
    # ========================================================
    for row_index, fila in enumerate(filas, start=5):
        for col_index, columna in enumerate(columnas, start=1):
            cell = ws.cell(row=row_index, column=col_index, value=_safe_value(fila.get(columna)))
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Congelar encabezado y activar autofiltro.
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(columnas))}{max(ws.max_row, 4)}"

    # Ancho automático simple por contenido.
    for col_index, columna in enumerate(columnas, start=1):
        letter = get_column_letter(col_index)
        max_length = len(columna) + 2
        for cell in ws[letter]:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)) + 2)
        ws.column_dimensions[letter].width = min(max_length, 45)

    wb.save(archivo)
    return archivo
