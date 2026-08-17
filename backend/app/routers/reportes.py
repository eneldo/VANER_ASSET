from datetime import date, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload

from app.core.auth_dependencies import require_roles
from app.database import get_db
from app.models.equipo import Equipo
from app.models.empresa import Empresa
from app.models.evidencia import Evidencia
from app.models.mantenimiento import Mantenimiento
from app.models.ot_incidencia import OtIncidencia
from app.models.ot_repuesto import OtRepuesto
from app.models.sede import Sede
from app.models.tecnico import Tecnico
from app.models.usuario import Usuario
from app.routers.evidencias import crear_url_firmada
from app.routers.reportes_publicados import _generar_pdf_ot
from app.services.evidencia_service import get_evidencia_path

router = APIRouter(prefix="/reportes", tags=["Reportes PRO"])
report_user = require_roles("ADMIN", "COORDINADOR")


def _rol(usuario: Usuario) -> str:
    return str(getattr(usuario, "rol", "") or "").upper()


def _empresa_permitida(usuario: Usuario, empresa_id: UUID | None = None) -> UUID | None:
    if _rol(usuario) == "ADMIN":
        return empresa_id

    empresa_usuario = getattr(usuario, "empresa_id", None)
    if not empresa_usuario:
        raise HTTPException(status_code=403, detail="El coordinador no tiene empresa asignada")
    if empresa_id and str(empresa_id) != str(empresa_usuario):
        raise HTTPException(status_code=403, detail="No puedes consultar reportes de otra empresa")
    return empresa_usuario


def _query_base(db: Session):
    return db.query(Mantenimiento).options(
        joinedload(Mantenimiento.equipo),
        joinedload(Mantenimiento.empresa),
        joinedload(Mantenimiento.sede),
        joinedload(Mantenimiento.tecnico).joinedload(Tecnico.usuario),
    )


def _aplicar_filtros(
    query,
    usuario: Usuario,
    empresa_id: UUID | None,
    sede_id: UUID | None,
    estado: str | None,
    fecha_inicio: date | None,
    fecha_fin: date | None,
):
    query = query.outerjoin(Equipo, Mantenimiento.equipo_id == Equipo.id)
    empresa_final = _empresa_permitida(usuario, empresa_id)

    if empresa_final:
        query = query.filter(
            or_(
                Mantenimiento.empresa_id == empresa_final,
                Equipo.empresa_id == empresa_final,
            )
        )
    if sede_id:
        query = query.filter(
            or_(
                Mantenimiento.sede_id == sede_id,
                Equipo.sede_id == sede_id,
            )
        )
    if estado:
        query = query.filter(Mantenimiento.estado == estado.strip().upper())
    if fecha_inicio:
        query = query.filter(Mantenimiento.fecha_programada >= datetime.combine(fecha_inicio, time.min))
    if fecha_fin:
        query = query.filter(Mantenimiento.fecha_programada < datetime.combine(fecha_fin + timedelta(days=1), time.min))

    return query


def _fecha(value):
    return value.isoformat() if value else None


def _contexto_mantenimiento(mantenimiento: Mantenimiento, db: Session):
    equipo = mantenimiento.equipo or db.query(Equipo).filter(Equipo.id == mantenimiento.equipo_id).first()
    empresa_id = mantenimiento.empresa_id or getattr(equipo, "empresa_id", None)
    sede_id = mantenimiento.sede_id or getattr(equipo, "sede_id", None)
    empresa = mantenimiento.empresa or (
        db.query(Empresa).filter(Empresa.id == empresa_id).first() if empresa_id else None
    )
    sede = mantenimiento.sede or (
        db.query(Sede).filter(Sede.id == sede_id).first() if sede_id else None
    )
    tecnico = mantenimiento.tecnico
    usuario_tecnico = getattr(tecnico, "usuario", None) if tecnico else None
    return equipo, empresa, sede, usuario_tecnico


def _descripciones_evidencias(evidencias):
    descripciones = {}
    for evidencia in evidencias:
        tipo = str(evidencia.tipo or "SOPORTE").upper()
        if tipo not in descripciones and str(evidencia.descripcion or "").strip():
            descripciones[tipo] = evidencia.descripcion.strip()
    return descripciones


def _serializar_mantenimiento(mantenimiento, db, evidencias=None):
    evidencias = evidencias or []
    equipo, empresa, sede, usuario_tecnico = _contexto_mantenimiento(mantenimiento, db)
    descripciones = _descripciones_evidencias(evidencias)
    fecha_finalizacion = mantenimiento.fecha_finalizacion or mantenimiento.fecha_fin

    return {
        "id": str(mantenimiento.id),
        "empresa_id": str(getattr(empresa, "id", "") or "") or None,
        "empresa": getattr(empresa, "nombre", None) or "Sin empresa",
        "sede_id": str(getattr(sede, "id", "") or "") or None,
        "sede": getattr(sede, "nombre", None) or "Sin sede",
        "equipo_id": str(getattr(equipo, "id", "") or "") or None,
        "equipo": getattr(equipo, "nombre", None) or "Sin equipo",
        "codigo_inventario": getattr(equipo, "inventario", None) or getattr(equipo, "codigo_id", None) or "",
        "inventario": getattr(equipo, "inventario", None),
        "codigo_id": getattr(equipo, "codigo_id", None),
        "serie": getattr(equipo, "serie", None),
        "marca": getattr(equipo, "marca", None),
        "modelo": getattr(equipo, "modelo", None),
        "ubicacion": getattr(equipo, "ubicacion", None),
        "tecnico": getattr(usuario_tecnico, "nombre_completo", None) or "Sin técnico",
        "tipo": mantenimiento.tipo or "",
        "estado": mantenimiento.estado or "",
        "fecha_programada": _fecha(mantenimiento.fecha_programada),
        "fecha_inicio": _fecha(mantenimiento.fecha_inicio),
        "fecha_fin": _fecha(fecha_finalizacion),
        "costo": float(mantenimiento.costo or 0),
        "descripcion": mantenimiento.descripcion or "",
        "estado_inicial": mantenimiento.estado_inicial or mantenimiento.estado_inicial_equipo or descripciones.get("ANTES", ""),
        "acciones_realizadas": mantenimiento.acciones_realizadas or descripciones.get("DURANTE", ""),
        "resultado_final": mantenimiento.resultado_final or descripciones.get("DESPUES", ""),
        "observaciones": mantenimiento.observaciones or "",
        "evidencias_total": len(evidencias),
    }


def consultar_mantenimientos(
    db: Session,
    usuario: Usuario,
    empresa_id: Optional[UUID] = None,
    sede_id: Optional[UUID] = None,
    estado: Optional[str] = None,
    fecha_inicio: Optional[date] = None,
    fecha_fin: Optional[date] = None,
):
    query = _aplicar_filtros(
        _query_base(db), usuario, empresa_id, sede_id, estado, fecha_inicio, fecha_fin
    )
    mantenimientos = query.order_by(
        Mantenimiento.fecha_programada.desc().nullslast(),
        Mantenimiento.created_at.desc(),
    ).all()
    ids = [item.id for item in mantenimientos]
    evidencias_map = {item_id: [] for item_id in ids}
    if ids:
        evidencias = db.query(Evidencia).filter(Evidencia.mantenimiento_id.in_(ids)).order_by(Evidencia.created_at).all()
        for evidencia in evidencias:
            evidencias_map.setdefault(evidencia.mantenimiento_id, []).append(evidencia)
    return [
        _serializar_mantenimiento(item, db, evidencias_map.get(item.id, []))
        for item in mantenimientos
    ]


def _obtener_mantenimiento_autorizado(db: Session, usuario: Usuario, mantenimiento_id: UUID):
    query = _aplicar_filtros(
        _query_base(db).filter(Mantenimiento.id == mantenimiento_id),
        usuario,
        None,
        None,
        None,
        None,
        None,
    )
    mantenimiento = query.first()
    if not mantenimiento:
        raise HTTPException(status_code=404, detail="Mantenimiento no encontrado")
    return mantenimiento


def _detalle_mantenimiento(db: Session, usuario: Usuario, mantenimiento_id: UUID):
    mantenimiento = _obtener_mantenimiento_autorizado(db, usuario, mantenimiento_id)
    evidencias = db.query(Evidencia).filter(
        Evidencia.mantenimiento_id == mantenimiento.id
    ).order_by(Evidencia.created_at).all()
    repuestos = db.query(OtRepuesto).filter(OtRepuesto.mantenimiento_id == mantenimiento.id).all()
    incidencias = db.query(OtIncidencia).filter(OtIncidencia.mantenimiento_id == mantenimiento.id).all()
    detalle = _serializar_mantenimiento(mantenimiento, db, evidencias)
    detalle["evidencias"] = [
        {
            "id": str(item.id),
            "tipo": item.tipo,
            "descripcion": item.descripcion,
            "nombre_original": item.nombre_original,
            "archivo_url": crear_url_firmada(item.id, item.archivo_url),
            "created_at": _fecha(item.created_at),
        }
        for item in evidencias
    ]
    detalle["repuestos"] = [
        {
            "descripcion": item.descripcion,
            "referencia": item.referencia,
            "cantidad": float(item.cantidad),
            "unidad": item.unidad,
        }
        for item in repuestos
    ]
    detalle["incidencias"] = [
        {
            "tipo": item.tipo,
            "severidad": item.severidad,
            "descripcion": item.descripcion,
            "resuelta": item.resuelta,
        }
        for item in incidencias
    ]
    return mantenimiento, detalle, evidencias


def _descarga(output: BytesIO, media_type: str, filename: str):
    output.seek(0)
    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _crear_excel_consolidado(data):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Mantenimientos"
    sheet.merge_cells("A1:Q1")
    sheet["A1"] = "REPORTE PRO DE MANTENIMIENTOS"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1E3A8A")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet["A2"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    headers = [
        "OT", "Empresa", "Sede", "Equipo", "Inventario", "Ubicación", "Técnico",
        "Tipo", "Estado", "Programado", "Inicio", "Finalización", "Estado inicial",
        "Acciones realizadas", "Resultado final", "Observaciones", "Evidencias",
    ]
    sheet.append([])
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="0F172A")
    border = Border(bottom=Side(style="thin", color="CBD5E1"))
    for cell in sheet[4]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for item in data:
        sheet.append([
            item["id"], item["empresa"], item["sede"], item["equipo"], item["codigo_inventario"],
            item["ubicacion"], item["tecnico"], item["tipo"], item["estado"], item["fecha_programada"],
            item["fecha_inicio"], item["fecha_fin"], item["estado_inicial"], item["acciones_realizadas"],
            item["resultado_final"], item["observaciones"], item["evidencias_total"],
        ])
    for row in sheet.iter_rows(min_row=5):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [38, 24, 24, 25, 18, 25, 24, 16, 16, 20, 20, 20, 32, 38, 32, 38, 12]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(4, index).column_letter].width = width
    output = BytesIO()
    workbook.save(output)
    return output


def _crear_excel_individual(detalle, evidencias):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Informe OT"
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "INFORME COMPLETO DE MANTENIMIENTO"
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1E3A8A")
    sheet["A1"].alignment = Alignment(horizontal="center")
    filas = [
        ("OT", detalle["id"]), ("Empresa", detalle["empresa"]), ("Sede", detalle["sede"]),
        ("Equipo", detalle["equipo"]), ("Inventario", detalle["codigo_inventario"]),
        ("Ubicación", detalle["ubicacion"]), ("Marca / Modelo", f'{detalle["marca"] or ""} / {detalle["modelo"] or ""}'),
        ("Serie", detalle["serie"]), ("Técnico", detalle["tecnico"]), ("Tipo", detalle["tipo"]),
        ("Estado", detalle["estado"]), ("Fecha programada", detalle["fecha_programada"]),
        ("Inicio", detalle["fecha_inicio"]), ("Finalización", detalle["fecha_fin"]),
        ("Estado inicial", detalle["estado_inicial"]), ("Acciones realizadas", detalle["acciones_realizadas"]),
        ("Resultado final", detalle["resultado_final"]), ("Observaciones", detalle["observaciones"]),
    ]
    for etiqueta, valor in filas:
        sheet.append([etiqueta, valor or "—"])
    for row in sheet.iter_rows(min_row=2, max_col=2):
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill("solid", fgColor="EFF6FF")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 80

    evidence_sheet = workbook.create_sheet("Evidencias")
    evidence_sheet.append(["Tipo", "Descripción", "Archivo", "Imagen"])
    for cell in evidence_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0F172A")
    for index, evidencia in enumerate(evidencias, start=2):
        evidence_sheet.cell(index, 1, evidencia.tipo or "SOPORTE")
        evidence_sheet.cell(index, 2, evidencia.descripcion or "")
        evidence_sheet.cell(index, 3, evidencia.nombre_original or Path(evidencia.archivo_url).name)
        path = get_evidencia_path(evidencia.archivo_url)
        if path.exists() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
            try:
                image = ExcelImage(str(path))
                image.width = 320
                image.height = 220
                evidence_sheet.add_image(image, f"D{index}")
                evidence_sheet.row_dimensions[index].height = 170
            except Exception:
                evidence_sheet.cell(index, 4, "Imagen no compatible")
    evidence_sheet.column_dimensions["A"].width = 16
    evidence_sheet.column_dimensions["B"].width = 55
    evidence_sheet.column_dimensions["C"].width = 35
    evidence_sheet.column_dimensions["D"].width = 48
    output = BytesIO()
    workbook.save(output)
    return output


@router.get("/filtros/empresas")
def obtener_empresas_filtro(
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(report_user),
):
    query = db.query(Empresa)
    empresa_id = _empresa_permitida(usuario)
    if empresa_id:
        query = query.filter(Empresa.id == empresa_id)
    return [{"id": str(item.id), "nombre": item.nombre} for item in query.order_by(Empresa.nombre).all()]


@router.get("/filtros/sedes")
def obtener_sedes_filtro(
    empresa_id: Optional[UUID] = Query(None),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(report_user),
):
    empresa_final = _empresa_permitida(usuario, empresa_id)
    query = db.query(Sede)
    if empresa_final:
        query = query.filter(Sede.empresa_id == empresa_final)
    return [
        {"id": str(item.id), "nombre": item.nombre, "empresa_id": str(item.empresa_id)}
        for item in query.order_by(Sede.nombre).all()
    ]


@router.get("/mantenimientos")
def reporte_mantenimientos(
    empresa_id: Optional[UUID] = Query(None),
    sede_id: Optional[UUID] = Query(None),
    estado: Optional[str] = Query(None),
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(report_user),
):
    data = consultar_mantenimientos(db, usuario, empresa_id, sede_id, estado, fecha_inicio, fecha_fin)
    return {"total": len(data), "items": data}


@router.get("/mantenimientos/excel")
def exportar_mantenimientos_excel(
    empresa_id: Optional[UUID] = Query(None),
    sede_id: Optional[UUID] = Query(None),
    estado: Optional[str] = Query(None),
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(report_user),
):
    data = consultar_mantenimientos(db, usuario, empresa_id, sede_id, estado, fecha_inicio, fecha_fin)
    return _descarga(
        _crear_excel_consolidado(data),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"reporte_mantenimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )


@router.get("/mantenimientos/pdf")
def exportar_mantenimientos_pdf(
    empresa_id: Optional[UUID] = Query(None),
    sede_id: Optional[UUID] = Query(None),
    estado: Optional[str] = Query(None),
    fecha_inicio: Optional[date] = Query(None),
    fecha_fin: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(report_user),
):
    data = consultar_mantenimientos(db, usuario, empresa_id, sede_id, estado, fecha_inicio, fecha_fin)
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet()
    rows = [["Empresa", "Sede", "Equipo", "Técnico", "Tipo", "Estado", "Programado", "Evidencias"]]
    rows.extend([
        [item["empresa"], item["sede"], item["equipo"], item["tecnico"], item["tipo"], item["estado"], item["fecha_programada"] or "", str(item["evidencias_total"])]
        for item in data
    ])
    table = Table(rows, repeatRows=1, colWidths=[100, 100, 125, 105, 70, 75, 105, 55])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
    ]))
    doc.build([
        Paragraph("REPORTE PRO DE MANTENIMIENTOS", styles["Title"]),
        Paragraph(f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles["Normal"]),
        Spacer(1, 12),
        table,
    ])
    return _descarga(output, "application/pdf", f"reporte_mantenimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf")


@router.get("/mantenimientos/{mantenimiento_id}/excel")
def exportar_mantenimiento_excel(
    mantenimiento_id: UUID,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(report_user),
):
    _, detalle, evidencias = _detalle_mantenimiento(db, usuario, mantenimiento_id)
    return _descarga(
        _crear_excel_individual(detalle, evidencias),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        f"informe_mantenimiento_{mantenimiento_id}.xlsx",
    )


@router.get("/mantenimientos/{mantenimiento_id}/pdf")
def exportar_mantenimiento_pdf(
    mantenimiento_id: UUID,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(report_user),
):
    mantenimiento = _obtener_mantenimiento_autorizado(db, usuario, mantenimiento_id)
    with TemporaryDirectory() as temp_dir:
        path = Path(temp_dir) / f"informe_mantenimiento_{mantenimiento_id}.pdf"
        _generar_pdf_ot(db, mantenimiento, path)
        output = BytesIO(path.read_bytes())
    return _descarga(output, "application/pdf", f"informe_mantenimiento_{mantenimiento_id}.pdf")


@router.get("/mantenimientos/{mantenimiento_id}")
def detalle_reporte_mantenimiento(
    mantenimiento_id: UUID,
    db: Session = Depends(get_db),
    usuario: Usuario = Depends(report_user),
):
    _, detalle, _ = _detalle_mantenimiento(db, usuario, mantenimiento_id)
    return detalle
