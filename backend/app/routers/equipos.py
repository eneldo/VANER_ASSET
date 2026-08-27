# =========================================================
# ROUTER EQUIPOS
# CRUD de equipos básicos - PASO 1 hoja de vida
# =========================================================

from datetime import datetime, timezone
from io import BytesIO
from uuid import UUID

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import pandas as pd
from sqlalchemy import func, or_
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.equipo import Equipo
from app.models.equipo_hoja_vida import EquipoHojaVida
from app.models.empresa import Empresa
from app.models.mantenimiento import Mantenimiento
from app.models.sede import Sede
from app.models.categoria import Categoria
from app.models.usuario import Usuario
from app.schemas.equipo import (
    EquipoCreate,
    EquipoUpdate,
    EquipoOut,
    EquipoAsignar,
    EquipoDevolver,
    EquipoTransferir,
    EquipoBaja,
    EquipoHistorialItem,
    EquipoHistorialOut,
    TIPOS_MOVIMIENTO,
    ESTADOS_EQUIPO,
)
from app.core.auth_dependencies import require_roles
from app.services.inventory_import_security import (
    read_inventory_upload,
    validate_inventory_shape,
)


router = APIRouter(
    prefix="/equipos",
    tags=["Equipos"],
    dependencies=[Depends(require_roles("ADMIN"))],
)


# Estados permitidos del equipo
ESTADOS_EQUIPO = [
    "OPERATIVO",
    "EN_MANTENIMIENTO",
    "FUERA_DE_SERVICIO",
    "BAJA"
]

# Criticidades permitidas del equipo
CRITICIDADES_EQUIPO = [
    "BAJA",
    "MEDIA",
    "ALTA",
    "CRITICA"
]


def normalizar_numero_inventario(value):
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    numero = str(value).strip()
    return numero or None


def validar_numero_inventario(
    db: Session,
    value,
    *,
    excluir_equipo_id: UUID | None = None,
):
    numero = normalizar_numero_inventario(value)
    if not numero:
        return None

    clave = numero.lower()
    filtros = [
        or_(
            func.lower(func.trim(Equipo.inventario)) == clave,
            func.lower(func.trim(Equipo.codigo_id)) == clave,
        )
    ]
    if excluir_equipo_id:
        filtros.append(Equipo.id != excluir_equipo_id)

    existente = db.query(Equipo).filter(*filtros).first()
    if existente:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Equipo ya existe: el número de inventario está registrado",
        )

    return numero

COLUMNAS_EXPORTACION = [
    "codigo_inventario",
    "nombre",
    "empresa",
    "sede",
    "categoria",
    "marca",
    "modelo",
    "serie",
    "ubicacion",
    "estado",
    "criticidad",
    "invima",
    "inventario",
    "activo",
    "fecha_creacion",
]


def _valor_excel(value, fallback="SIN DATO"):
    """Convierte valores a texto seguro y evita formulas inyectadas en Excel."""
    if value is None or str(value).strip() == "":
        return fallback
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    text = str(value).strip()
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def crear_excel_inventario(filas):
    """Genera un libro compatible con importacion y datos de respaldo."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Inventario"

    header_fill = PatternFill("solid", fgColor="17365D")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")

    for column_index, column_name in enumerate(COLUMNAS_EXPORTACION, start=1):
        cell = sheet.cell(row=1, column=column_index, value=column_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    for row_index, fila in enumerate(filas, start=2):
        for column_index, column_name in enumerate(COLUMNAS_EXPORTACION, start=1):
            fallback = (
                f"SIN-INVENTARIO-{row_index:04d}"
                if column_name == "codigo_inventario"
                else "SIN DATO"
            )
            value = fila.get(column_name)
            if column_name == "codigo_inventario":
                value = (
                    normalizar_numero_inventario(value)
                    or normalizar_numero_inventario(fila.get("inventario"))
                )
            sheet.cell(
                row=row_index,
                column=column_index,
                value=_valor_excel(value, fallback),
            ).alignment = Alignment(vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(COLUMNAS_EXPORTACION))}{max(sheet.max_row, 1)}"
    )
    sheet.row_dimensions[1].height = 28

    widths = [22, 24, 26, 28, 34, 18, 22, 20, 36, 24, 16, 18, 20, 12, 22]
    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width

    estado_validation = DataValidation(
        type="list",
        formula1='"OPERATIVO,EN_MANTENIMIENTO,FUERA_DE_SERVICIO,BAJA"',
        allow_blank=False,
    )
    criticidad_validation = DataValidation(
        type="list",
        formula1='"BAJA,MEDIA,ALTA,CRITICA"',
        allow_blank=False,
    )
    sheet.add_data_validation(estado_validation)
    sheet.add_data_validation(criticidad_validation)
    estado_validation.add(f"J2:J{max(sheet.max_row, 2)}")
    criticidad_validation.add(f"K2:K{max(sheet.max_row, 2)}")

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def validar_relaciones_equipo(data, db: Session):
    """
    Valida que empresa, sede y categoría existan.
    También valida que la sede pertenezca a la empresa.
    """

    # Validar empresa
    empresa = db.query(Empresa).filter(Empresa.id == data.empresa_id).first()

    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="La empresa asociada no existe"
        )

    # Validar sede
    sede = db.query(Sede).filter(Sede.id == data.sede_id).first()

    if not sede:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="La sede asociada no existe"
        )

    # Validar que la sede pertenezca a la empresa seleccionada
    if sede.empresa_id != data.empresa_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La sede no pertenece a la empresa seleccionada"
        )

    # Validar categoría si se envió
    if data.categoria_id:
        categoria = db.query(Categoria).filter(
            Categoria.id == data.categoria_id
        ).first()

        if not categoria or not categoria.activo:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="La categoría asociada no existe"
            )


def validar_responsable(db: Session, responsable_id: UUID | None, empresa_id: UUID):
    if responsable_id is None:
        return None
    responsable = db.query(Usuario).filter(Usuario.id == responsable_id).first()
    if not responsable or not responsable.activo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Responsable no encontrado o inactivo",
        )
    if responsable.empresa_id is not None and responsable.empresa_id != empresa_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El responsable no pertenece a la empresa del equipo",
        )
    return responsable


def validar_equipo_movible(equipo: Equipo):
    if not equipo.activo or equipo.estado == "BAJA":
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="No se permiten movimientos sobre equipos inactivos o dados de baja",
        )


def validar_estado_y_criticidad(estado: str, criticidad: str):
    """
    Valida valores permitidos de estado y criticidad.
    """

    if estado not in ESTADOS_EQUIPO:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Estado no permitido. Use uno de: {ESTADOS_EQUIPO}"
        )

    if criticidad not in CRITICIDADES_EQUIPO:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Criticidad no permitida. Use una de: {CRITICIDADES_EQUIPO}"
        )


@router.post("/", response_model=EquipoOut)
def crear_equipo(data: EquipoCreate, db: Session = Depends(get_db)):
    """
    Crea un equipo con datos básicos.
    Este es el PASO 1 del flujo de hoja de vida en dos pasos.
    """

    # Validar relaciones empresa/sede/categoría
    validar_relaciones_equipo(data, db)

    # Validar estado y criticidad
    validar_estado_y_criticidad(data.estado, data.criticidad)
    validar_responsable(db, data.responsable_id, data.empresa_id)

    # Validar código único si se envía
    if data.codigo_id:
        existente = db.query(Equipo).filter(
            Equipo.codigo_id == data.codigo_id
        ).first()

        if existente:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe un equipo con ese Código ID"
            )

    datos = data.model_dump()
    datos["inventario"] = validar_numero_inventario(db, datos.get("inventario"))
    nuevo_equipo = Equipo(**datos)

    db.add(nuevo_equipo)
    db.commit()
    db.refresh(nuevo_equipo)

    return nuevo_equipo


@router.get("/", response_model=list[EquipoOut])
def listar_equipos(
    db: Session = Depends(get_db),
    estado: str | None = None,
    criticidad: str | None = None,
    categoria_id: UUID | None = None,
    sede_id: UUID | None = None,
    responsable_id: UUID | None = None,
    q: str | None = None,
    solo_activos: bool = True,
):
    """
    Lista todos los equipos registrados con filtros opcionales.

    Filtros disponibles:
    - estado: OPERATIVO, EN_MANTENIMIENTO, FUERA_DE_SERVICIO, BAJA
    - criticidad: BAJA, MEDIA, ALTA, CRITICA
    - categoria_id: UUID de categoría
    - sede_id: UUID de sede
    - responsable_id: UUID de responsable
    - q: búsqueda por nombre, serie, inventario, codigo_id
    - solo_activos: true/false (default true)
    """
    query = db.query(Equipo)

    if solo_activos:
        query = query.filter(Equipo.activo == True)

    if estado:
        if estado not in ESTADOS_EQUIPO:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Estado no permitido. Use uno de: {ESTADOS_EQUIPO}"
            )
        query = query.filter(Equipo.estado == estado)

    if criticidad:
        if criticidad not in CRITICIDADES_EQUIPO:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Criticidad no permitida. Use una de: {CRITICIDADES_EQUIPO}"
            )
        query = query.filter(Equipo.criticidad == criticidad)

    if categoria_id:
        query = query.filter(Equipo.categoria_id == categoria_id)

    if sede_id:
        query = query.filter(Equipo.sede_id == sede_id)

    if responsable_id:
        query = query.filter(Equipo.responsable_id == responsable_id)

    if q:
        search = f"%{q}%"
        query = query.filter(
            or_(
                Equipo.nombre.ilike(search),
                Equipo.serie.ilike(search),
                Equipo.inventario.ilike(search),
                Equipo.codigo_id.ilike(search),
            )
        )

    equipos = query.order_by(Equipo.created_at.desc()).all()
    return equipos


@router.get("/empresa/{empresa_id}", response_model=list[EquipoOut])
def listar_equipos_por_empresa(empresa_id: UUID, db: Session = Depends(get_db)):
    """
    Lista equipos filtrados por empresa.
    Esta ruta será usada por el portal EMPRESA.
    """

    equipos = db.query(Equipo).filter(
        Equipo.empresa_id == empresa_id
    ).order_by(Equipo.created_at.desc()).all()

    return equipos


@router.get("/sede/{sede_id}", response_model=list[EquipoOut])
def listar_equipos_por_sede(sede_id: UUID, db: Session = Depends(get_db)):
    """
    Lista equipos filtrados por sede.
    """

    equipos = db.query(Equipo).filter(
        Equipo.sede_id == sede_id
    ).order_by(Equipo.created_at.desc()).all()

    return equipos


@router.get("/exportar")
def exportar_inventario_equipos(db: Session = Depends(get_db)):
    """Descarga el inventario completo en un archivo Excel."""
    registros = (
        db.query(
            Equipo,
            Empresa.nombre.label("empresa_nombre"),
            Sede.nombre.label("sede_nombre"),
            Categoria.nombre.label("categoria_nombre"),
        )
        .outerjoin(Empresa, Empresa.id == Equipo.empresa_id)
        .outerjoin(Sede, Sede.id == Equipo.sede_id)
        .outerjoin(Categoria, Categoria.id == Equipo.categoria_id)
        .order_by(Empresa.nombre, Sede.nombre, Equipo.nombre)
        .all()
    )

    filas = [
        {
            "codigo_inventario": equipo.codigo_id,
            "nombre": equipo.nombre,
            "empresa": empresa_nombre,
            "sede": sede_nombre,
            "categoria": categoria_nombre,
            "marca": equipo.marca,
            "modelo": equipo.modelo,
            "serie": equipo.serie,
            "ubicacion": equipo.ubicacion,
            "estado": equipo.estado,
            "criticidad": equipo.criticidad,
            "invima": equipo.invima,
            "inventario": equipo.inventario,
            "activo": "SI" if equipo.activo else "NO",
            "fecha_creacion": equipo.created_at,
        }
        for equipo, empresa_nombre, sede_nombre, categoria_nombre in registros
    ]

    filename = f"inventario_equipos_vaner_asset_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        crear_excel_inventario(filas),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{equipo_id}", response_model=EquipoOut)
def obtener_equipo(equipo_id: UUID, db: Session = Depends(get_db)):
    """
    Obtiene un equipo por ID.
    """

    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()

    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )

    return equipo


@router.put("/{equipo_id}", response_model=EquipoOut)
def actualizar_equipo(
    equipo_id: UUID,
    data: EquipoUpdate,
    db: Session = Depends(get_db)
):
    """
    Actualiza datos básicos de un equipo.
    """

    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()

    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )

    datos = data.model_dump(exclude_unset=True)

    if "inventario" in datos:
        inventario_actual = normalizar_numero_inventario(equipo.inventario)
        inventario_nuevo = normalizar_numero_inventario(datos["inventario"])
        datos["inventario"] = (
            inventario_nuevo
            if inventario_nuevo == inventario_actual
            else validar_numero_inventario(
                db,
                inventario_nuevo,
                excluir_equipo_id=equipo_id,
            )
        )

    # Validar código único si cambia
    if "codigo_id" in datos and datos["codigo_id"]:
        existente = db.query(Equipo).filter(
            Equipo.codigo_id == datos["codigo_id"],
            Equipo.id != equipo_id
        ).first()

        if existente:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Ya existe otro equipo con ese Código ID"
            )

    # Validar estado si se envía
    if "estado" in datos and datos["estado"] not in ESTADOS_EQUIPO:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Estado no permitido. Use uno de: {ESTADOS_EQUIPO}"
        )

    # Validar criticidad si se envía
    if "criticidad" in datos and datos["criticidad"] not in CRITICIDADES_EQUIPO:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Criticidad no permitida. Use una de: {CRITICIDADES_EQUIPO}"
        )

    # Si se cambia empresa/sede, validar consistencia
    empresa_id_final = datos.get("empresa_id", equipo.empresa_id)
    sede_id_final = datos.get("sede_id", equipo.sede_id)

    empresa = db.query(Empresa).filter(Empresa.id == empresa_id_final).first()

    if not empresa:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="La empresa asociada no existe"
        )

    sede = db.query(Sede).filter(Sede.id == sede_id_final).first()

    if not sede:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="La sede asociada no existe"
        )

    if sede.empresa_id != empresa_id_final:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="La sede no pertenece a la empresa seleccionada"
        )

    validar_responsable(
        db,
        datos.get("responsable_id", equipo.responsable_id),
        empresa_id_final,
    )

    # Validar categoría si se cambia
    if "categoria_id" in datos and datos["categoria_id"]:
        categoria = db.query(Categoria).filter(
            Categoria.id == datos["categoria_id"]
        ).first()

        if not categoria or not categoria.activo:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="La categoría asociada no existe"
            )

    # Aplicar cambios
    for campo, valor in datos.items():
        setattr(equipo, campo, valor)

    db.commit()
    db.refresh(equipo)

    return equipo


@router.delete("/{equipo_id}")
def eliminar_equipo(equipo_id: UUID, db: Session = Depends(get_db)):
    """
    Elimina un equipo que todavía no tiene historial de mantenimiento.
    """

    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()

    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )

    mantenimiento = db.query(Mantenimiento.id).filter(
        Mantenimiento.equipo_id == equipo_id
    ).first()
    if mantenimiento:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "No se puede eliminar el equipo porque tiene mantenimientos "
                "asociados. Puede marcarlo como inactivo o en estado de baja."
            ),
        )

    try:
        db.query(EquipoHojaVida).filter(
            EquipoHojaVida.equipo_id == equipo_id
        ).delete(synchronize_session=False)
        db.delete(equipo)
        db.commit()
    except IntegrityError as error:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=(
                "No se puede eliminar el equipo porque tiene información "
                "operativa asociada."
            ),
        ) from error

    return {"message": "Equipo eliminado correctamente"}

# =========================================================
# CONTROL DE MOVIMIENTOS EQUIPO - FASE 7
# =========================================================

TIPOS_MOVIMIENTO_VALIDOS = [
    "ASIGNACION",
    "DEVOLUCION",
    "TRANSFERENCIA",
    "BAJA"
]

def _agregar_historial(equipo, campo, anterior, nuevo, usuario_id=None, tipo_movimiento=None, observacion=None, timestamp=None):
    entrada = {
        "timestamp": (timestamp or datetime.now(timezone.utc)).isoformat(),
        "campo": campo,
        "anterior": str(anterior) if anterior is not None else None,
        "nuevo": str(nuevo) if nuevo is not None else None,
        "usuario_id": str(usuario_id) if usuario_id is not None else None,
        "tipo_movimiento": tipo_movimiento,
        "observacion": observacion,
    }
    equipo.historial_cambios = [*(equipo.historial_cambios or []), entrada]


def _guardar_movimiento(db: Session, equipo: Equipo):
    try:
        db.commit()
        db.refresh(equipo)
    except Exception:
        db.rollback()
        raise
    return equipo


@router.post("/{equipo_id}/asignar", response_model=EquipoOut)
def asignar_equipo(
    equipo_id: UUID,
    data: EquipoAsignar,
    db: Session = Depends(get_db),
    usuario_actual: Usuario = Depends(require_roles("ADMIN")),
):
    """
    Asigna un equipo a un responsable.
    Actualiza responsable_id, ubicacion (opcional), estado a OPERATIVO,
    y registra en historial_cambios.
    """
    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()

    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )

    validar_equipo_movible(equipo)
    validar_responsable(db, data.responsable_id, equipo.empresa_id)

    # Guardar valores anteriores para historial
    responsable_anterior = equipo.responsable_id
    ubicacion_anterior = equipo.ubicacion
    estado_anterior = equipo.estado

    # Aplicar cambios
    equipo.responsable_id = data.responsable_id
    if data.ubicacion:
        equipo.ubicacion = data.ubicacion
    equipo.estado = "OPERATIVO"

    # Registrar en historial
    _agregar_historial(equipo, "responsable_id", responsable_anterior, data.responsable_id, usuario_actual.id, "ASIGNACION", data.observacion)
    if data.ubicacion:
        _agregar_historial(equipo, "ubicacion", ubicacion_anterior, data.ubicacion, usuario_actual.id, "ASIGNACION", data.observacion)
    if estado_anterior != "OPERATIVO":
        _agregar_historial(equipo, "estado", estado_anterior, "OPERATIVO", usuario_actual.id, "ASIGNACION", data.observacion)

    return _guardar_movimiento(db, equipo)


@router.post("/{equipo_id}/devolver", response_model=EquipoOut)
def devolver_equipo(
    equipo_id: UUID,
    data: EquipoDevolver,
    db: Session = Depends(get_db),
    usuario_actual: Usuario = Depends(require_roles("ADMIN")),
):
    """
    Devuelve un equipo (libera responsable).
    Pone responsable_id = None, estado = OPERATIVO (o FUERA_DE_SERVICIO si estaba en mantenimiento),
    y registra en historial_cambios.
    """
    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()

    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )

    validar_equipo_movible(equipo)
    if not equipo.responsable_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El equipo no tiene responsable asignado"
        )

    # Guardar valores anteriores
    responsable_anterior = equipo.responsable_id
    ubicacion_anterior = equipo.ubicacion
    estado_anterior = equipo.estado

    # Aplicar cambios
    equipo.responsable_id = None
    if data.ubicacion:
        equipo.ubicacion = data.ubicacion

    # Determinar nuevo estado
    if estado_anterior == "EN_MANTENIMIENTO":
        equipo.estado = "FUERA_DE_SERVICIO"
    else:
        equipo.estado = "OPERATIVO"

    # Registrar en historial
    _agregar_historial(equipo, "responsable_id", responsable_anterior, None, usuario_actual.id, "DEVOLUCION", data.observacion)
    if data.ubicacion:
        _agregar_historial(equipo, "ubicacion", ubicacion_anterior, data.ubicacion, usuario_actual.id, "DEVOLUCION", data.observacion)
    if estado_anterior != equipo.estado:
        _agregar_historial(equipo, "estado", estado_anterior, equipo.estado, usuario_actual.id, "DEVOLUCION", data.observacion)

    return _guardar_movimiento(db, equipo)


@router.post("/{equipo_id}/transferir", response_model=EquipoOut)
def transferir_equipo(
    equipo_id: UUID,
    data: EquipoTransferir,
    db: Session = Depends(get_db),
    usuario_actual: Usuario = Depends(require_roles("ADMIN")),
):
    """
    Transfiere un equipo entre sedes/áreas/responsables.
    Actualiza campos según lo enviado y registra en historial_cambios.
    """
    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()

    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )

    validar_equipo_movible(equipo)
    cambios = []

    # Validar y aplicar nuevo responsable
    if data.nuevo_responsable_id is not None:
        if data.nuevo_responsable_id:
            validar_responsable(db, data.nuevo_responsable_id, equipo.empresa_id)
        cambios.append(("responsable_id", equipo.responsable_id, data.nuevo_responsable_id))
        equipo.responsable_id = data.nuevo_responsable_id

    # Validar y aplicar nueva sede
    if data.nueva_sede_id is not None:
        from app.models.sede import Sede
        sede = db.query(Sede).filter(Sede.id == data.nueva_sede_id).first()
        if not sede:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Nueva sede no encontrada"
            )
        # Validar que la sede pertenezca a la misma empresa
        if sede.empresa_id != equipo.empresa_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La sede no pertenece a la misma empresa del equipo"
            )
        cambios.append(("sede_id", equipo.sede_id, data.nueva_sede_id))
        equipo.sede_id = data.nueva_sede_id

    # Aplicar nueva ubicación
    if data.nueva_ubicacion is not None:
        cambios.append(("ubicacion", equipo.ubicacion, data.nueva_ubicacion))
        equipo.ubicacion = data.nueva_ubicacion

    if not cambios:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Debe enviar al menos un campo para transferir"
        )

    # Registrar en historial
    for campo, anterior, nuevo in cambios:
        _agregar_historial(equipo, campo, anterior, nuevo, usuario_actual.id, "TRANSFERENCIA", data.observacion)

    return _guardar_movimiento(db, equipo)


@router.post("/{equipo_id}/baja", response_model=EquipoOut)
def dar_baja_equipo(
    equipo_id: UUID,
    data: EquipoBaja,
    db: Session = Depends(get_db),
    usuario_actual: Usuario = Depends(require_roles("ADMIN")),
):
    """
    Da de baja un equipo.
    Cambia estado a BAJA o FUERA_DE_SERVICIO, libera responsable,
    y registra en historial_cambios.
    """
    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()

    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )

    validar_equipo_movible(equipo)

    # Validar estado final
    if data.estado_final not in ["BAJA", "FUERA_DE_SERVICIO"]:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Estado final debe ser BAJA o FUERA_DE_SERVICIO"
        )

    # Guardar valores anteriores
    estado_anterior = equipo.estado
    responsable_anterior = equipo.responsable_id

    # Aplicar cambios
    equipo.estado = data.estado_final
    equipo.activo = data.estado_final != "BAJA"
    equipo.responsable_id = None

    # Registrar en historial
    _agregar_historial(equipo, "estado", estado_anterior, data.estado_final, usuario_actual.id, "BAJA", data.motivo)
    if responsable_anterior:
        _agregar_historial(equipo, "responsable_id", responsable_anterior, None, usuario_actual.id, "BAJA", data.motivo)

    return _guardar_movimiento(db, equipo)


@router.get("/{equipo_id}/historial", response_model=EquipoHistorialOut)
def obtener_historial_equipo(
    equipo_id: UUID,
    db: Session = Depends(get_db)
):
    """
    Obtiene el historial completo de movimientos de un equipo.
    """
    equipo = db.query(Equipo).filter(Equipo.id == equipo_id).first()

    if not equipo:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Equipo no encontrado"
        )

    # Convertir historial_cambios JSON a lista de EquipoHistorialItem
    historial_items = []
    if equipo.historial_cambios:
        for item in equipo.historial_cambios:
            historial_items.append(EquipoHistorialItem(
                timestamp=datetime.fromisoformat(item["timestamp"]) if isinstance(item["timestamp"], str) else item["timestamp"],
                campo=item.get("campo", ""),
                anterior=item.get("anterior"),
                nuevo=item.get("nuevo"),
                usuario_id=str(item["usuario_id"]) if item.get("usuario_id") is not None else None,
                tipo_movimiento=item.get("tipo_movimiento"),
                observacion=item.get("observacion"),
            ))

    return EquipoHistorialOut(
        id=equipo.id,
        nombre=equipo.nombre,
        estado=equipo.estado,
        responsable_id=equipo.responsable_id,
        ubicacion=equipo.ubicacion,
        vida_util_meses=equipo.vida_util_meses,
        historial_cambios=historial_items
    )

# =========================================================
# IMPORTAR INVENTARIO DESDE EXCEL / CSV
# =========================================================


def normalizar_celda_importacion(value):
    if value is None:
        return None

    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    texto = str(value).strip()
    return texto or None

def requerir_celda_importacion(row, columna, etiqueta):
    value = normalizar_celda_importacion(row.get(columna))
    if not value:
        raise ValueError(f"{etiqueta} es obligatorio")
    return value

def mensaje_error_importacion(error):
    if isinstance(error, HTTPException):
        return str(error.detail)
    if isinstance(error, IntegrityError):
        return "El código de inventario o el número de inventario ya está registrado"
    return str(error) or "Error inesperado procesando la fila"

@router.post("/importar")
async def importar_equipos(
    archivo: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Importa equipos y aísla los errores para que una fila no aborte todo el archivo."""

    try:
        contenido, extension = await read_inventory_upload(archivo)

        if extension == ".csv":
            df = pd.read_csv(BytesIO(contenido), encoding="utf-8-sig")
        else:
            df = pd.read_excel(BytesIO(contenido), engine="openpyxl")
        validate_inventory_shape(len(df.index), len(df.columns))
    except HTTPException:
        raise
    except Exception as error:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Error leyendo archivo: {error}",
        ) from error

    df.columns = [str(column).strip().lower() for column in df.columns]
    columnas = [
        "codigo_inventario",
        "nombre",
        "empresa",
        "sede",
        "categoria",
        "marca",
        "modelo",
        "serie",
        "ubicacion",
        "estado",
        "criticidad",
    ]
    faltantes = [column for column in columnas if column not in df.columns]

    if faltantes:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Faltan columnas: {faltantes}",
        )

    creados = 0
    errores = []
    codigos_archivo = set()
    inventarios_archivo = set()

    for index, row in df.iterrows():
        try:
            with db.begin_nested():
                codigo_inventario = requerir_celda_importacion(
                    row,
                    "codigo_inventario",
                    "Código de inventario",
                )
                clave_codigo = codigo_inventario.lower()
                if clave_codigo in codigos_archivo:
                    raise ValueError(
                        "Equipo ya existe: código de inventario repetido en el archivo"
                    )

                codigo_existente = db.query(Equipo).filter(
                    func.lower(func.trim(Equipo.codigo_id)) == clave_codigo
                ).first()
                if codigo_existente:
                    raise ValueError(
                        "Equipo ya existe: el código de inventario está registrado"
                    )

                nombre = requerir_celda_importacion(
                    row,
                    "nombre",
                    "Nombre del equipo",
                )
                empresa_nombre = requerir_celda_importacion(
                    row,
                    "empresa",
                    "Empresa",
                )
                sede_nombre = requerir_celda_importacion(row, "sede", "Sede")
                categoria_nombre = requerir_celda_importacion(
                    row,
                    "categoria",
                    "Categoría",
                )

                empresa = db.query(Empresa).filter(
                    Empresa.nombre.ilike(empresa_nombre)
                ).first()
                if not empresa:
                    raise ValueError("Empresa no encontrada")

                sede = db.query(Sede).filter(
                    Sede.nombre.ilike(sede_nombre),
                    Sede.empresa_id == empresa.id,
                ).first()
                if not sede:
                    raise ValueError("Sede no encontrada para la empresa indicada")

                categoria = db.query(Categoria).filter(
                    Categoria.nombre.ilike(categoria_nombre),
                    Categoria.activo.is_(True),
                ).first()
                if not categoria:
                    raise ValueError("Categoría no encontrada o no permitida")

                estado = requerir_celda_importacion(
                    row,
                    "estado",
                    "Estado",
                ).upper()
                criticidad = requerir_celda_importacion(
                    row,
                    "criticidad",
                    "Criticidad",
                ).upper()
                validar_estado_y_criticidad(estado, criticidad)

                numero_inventario = normalizar_celda_importacion(
                    row.get("inventario") if "inventario" in df.columns else None
                )
                if numero_inventario and numero_inventario.upper() in {
                    "SIN DATO",
                    "N/A",
                    "NA",
                }:
                    numero_inventario = None

                if numero_inventario:
                    clave_inventario = numero_inventario.lower()
                    if clave_inventario in inventarios_archivo:
                        raise ValueError(
                            "Equipo ya existe: número de inventario repetido en el archivo"
                        )
                    numero_inventario = validar_numero_inventario(
                        db,
                        numero_inventario,
                    )

                nuevo = Equipo(
                    nombre=nombre,
                    empresa_id=empresa.id,
                    sede_id=sede.id,
                    categoria_id=categoria.id,
                    marca=normalizar_celda_importacion(row.get("marca")),
                    modelo=normalizar_celda_importacion(row.get("modelo")),
                    serie=normalizar_celda_importacion(row.get("serie")),
                    ubicacion=normalizar_celda_importacion(row.get("ubicacion")),
                    codigo_id=codigo_inventario,
                    inventario=numero_inventario,
                    estado=estado,
                    criticidad=criticidad,
                    activo=True,
                )
                db.add(nuevo)
                db.flush()

            creados += 1
            codigos_archivo.add(clave_codigo)
            if numero_inventario:
                inventarios_archivo.add(numero_inventario.lower())

        except Exception as error:
            errores.append({
                "fila": int(index + 2),
                "error": mensaje_error_importacion(error),
            })

    db.commit()

    return {
        "creados": creados,
        "omitidos": len(errores),
        "errores": errores,
    }
